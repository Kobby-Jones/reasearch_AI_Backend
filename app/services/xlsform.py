"""Export an instrument to XLSForm (the survey/choices/settings workbook used by
KoboToolbox and ODK Collect). Lets researchers build here, then field offline.

Type mapping:
  single_choice / dropdown / yes_no -> select_one <list>
  multiple_choice                   -> select_multiple <list>
  likert / rating                   -> select_one <list>  (numeric values 1..N)
  short_text                        -> text
  long_text                         -> text (appearance: multiline)
  numeric                           -> integer
  date                              -> date
"""
from __future__ import annotations

import re

from openpyxl import Workbook
from openpyxl.styles import Font

_HEADER = Font(name="Arial", bold=True)
_BODY = Font(name="Arial")


def _slug(text: str, fallback: str) -> str:
    s = re.sub(r"[^a-z0-9_]+", "_", (text or "").strip().lower()).strip("_")
    if not s or s[0].isdigit():
        s = f"q_{s}" if s else fallback
    return s[:40]


def _unique(name: str, used: set[str]) -> str:
    base, i = name, 1
    while name in used:
        name = f"{base}_{i}"
        i += 1
    used.add(name)
    return name


def build_xlsform(structure: dict, title: str, out_path: str) -> str:
    wb = Workbook()
    survey = wb.active
    survey.title = "survey"
    choices = wb.create_sheet("choices")
    settings = wb.create_sheet("settings")

    survey.append(["type", "name", "label", "required", "hint", "appearance"])
    choices.append(["list_name", "name", "label"])
    settings.append(["form_title", "form_id"])
    settings.append([title or "Research Survey", _slug(title, "research_form")])

    used_names: set[str] = set()
    used_lists: set[str] = set()

    def add_choice_list(list_name: str, pairs: list[tuple[str, str]]) -> None:
        for name, label in pairs:
            choices.append([list_name, name, label])

    for s_idx, section in enumerate(structure.get("sections") or []):
        sec_name = _unique(_slug(section.get("title", ""), f"section_{s_idx+1}"), used_names)
        survey.append(["begin group", sec_name, section.get("title", f"Section {s_idx+1}"), "", "", "field-list"])

        for i_idx, item in enumerate(section.get("items") or []):
            if not isinstance(item, dict):
                item = {"text": str(item), "type": "short_text"}
            qtype = item.get("type") or "likert"
            label = item.get("text", "")
            name = _unique(_slug(item.get("id") or label, f"q{s_idx+1}_{i_idx+1}"), used_names)
            required = "yes" if item.get("required", True) else ""
            appearance = ""
            xls_type = "text"

            if qtype in ("single_choice", "dropdown", "yes_no"):
                opts = item.get("options") or (["Yes", "No"] if qtype == "yes_no" else [])
                list_name = _unique(f"{name}_opts", used_lists)
                add_choice_list(list_name, [(_slug(o, f"opt{k+1}"), o) for k, o in enumerate(opts)])
                xls_type = f"select_one {list_name}"
                if qtype == "dropdown":
                    appearance = "minimal"
            elif qtype == "multiple_choice":
                opts = item.get("options") or []
                list_name = _unique(f"{name}_opts", used_lists)
                add_choice_list(list_name, [(_slug(o, f"opt{k+1}"), o) for k, o in enumerate(opts)])
                xls_type = f"select_multiple {list_name}"
            elif qtype in ("likert", "rating"):
                labels = item.get("scale_labels") or ["1", "2", "3", "4", "5"]
                list_name = _unique(f"{name}_scale", used_lists)
                # numeric values keep responses analysis-friendly
                add_choice_list(list_name, [(str(k + 1), lab) for k, lab in enumerate(labels)])
                xls_type = f"select_one {list_name}"
                if qtype == "rating":
                    appearance = "likert"
            elif qtype == "numeric":
                xls_type = "integer"
            elif qtype == "date":
                xls_type = "date"
            elif qtype == "long_text":
                xls_type = "text"
                appearance = "multiline"
            else:
                xls_type = "text"

            survey.append([xls_type, name, label, required, "", appearance])

        survey.append(["end group", "", "", "", "", ""])

    # styling: header row bold + sensible widths
    for ws in (survey, choices, settings):
        for cell in ws[1]:
            cell.font = _HEADER
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = _BODY
    for col, w in {"A": 22, "B": 26, "C": 60, "D": 10, "E": 24, "F": 14}.items():
        survey.column_dimensions[col].width = w
    for col, w in {"A": 26, "B": 22, "C": 44}.items():
        choices.column_dimensions[col].width = w

    wb.save(out_path)
    return out_path
